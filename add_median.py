#!/usr/bin/env python3
"""
add_median.py  -  Add a median fragment-size column to a TapeStation sample table.

WHAT IT DOES
  Reads the electropherogram export (the raw trace) plus the sample table from a
  TapeStation / Fragment Analyzer run, computes the MEDIAN fragment size for every
  sample (the 50%-cumulative-area point of the size distribution), and writes a new
  Excel file that is the sample table with an extra "Median Size [bp]" column.

HOW TO RUN
  1) Put this file in the folder that contains your exported CSVs, then:
         python add_median.py
  2) Or point it at the folder:
         python add_median.py "C:/path/to/exported/csvs"
  3) Or pass the two files explicitly:
         python add_median.py run_Electropherogram.csv run_sampleTable.csv

  Output:  <run-prefix>_withMedian.xlsx   (next to the inputs)
  Extra stats (mode / mean / Q25 / Q75):  add  --full

REQUIREMENTS
  pip install pandas numpy scipy openpyxl
"""
import sys, glob, os
import numpy as np
import pandas as pd
from scipy.signal import find_peaks
from scipy.interpolate import PchipInterpolator
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

LADDER_SIZES = np.array([25, 50, 100, 200, 300, 400, 500, 700, 1000, 1500])


# ----------------------------- median engine -------------------------------- #

def _read_table(path):
    if str(path).lower().endswith((".xlsx", ".xls")):
        return pd.read_excel(path)
    return pd.read_csv(path, encoding="latin-1")


def _calibration(ladder):
    pk, _ = find_peaks(ladder, height=40, distance=5)
    if len(pk) != len(LADDER_SIZES):
        # fall back: take the strongest 10 peaks
        pk, props = find_peaks(ladder, distance=5)
        order = np.argsort(props["peak_heights"] if "peak_heights" in props
                           else ladder[pk])[-len(LADDER_SIZES):]
        pk = np.sort(pk[order])
    if len(pk) != len(LADDER_SIZES):
        raise ValueError(f"Could not identify the 10 ladder peaks (found {len(pk)}).")
    return PchipInterpolator(pk, np.log10(LADDER_SIZES)), pk[0], pk[-1]


LOWER_BP = 200
UPPER_BP = 1200


def _stats(sig, calib, LM_l, UM_l):
    sig = sig.astype(float)
    rows = np.arange(len(sig))
    p, _ = find_peaks(sig, height=max(sig.max() * 0.15, 5), distance=5)
    if len(p) < 2:
        return None
    lm, um = p[0], p[-1]                       # 25 bp and 1500 bp markers
    size = 10 ** calib(LM_l + (UM_l - LM_l) * (rows - lm) / (um - lm))
    mask = (size >= LOWER_BP) & (size <= UPPER_BP)
    if mask.sum() < 2:
        return None
    w = np.clip(sig, 0, None)
    sz, ww = size[mask], w[mask]
    if ww.sum() <= 0:
        return None
    o = np.argsort(sz); sz, ww = sz[o], ww[o]
    cum = np.cumsum(ww) / ww.sum()
    return dict(median=float(np.interp(0.5, cum, sz)),
                mode=float(sz[np.argmax(ww)]),
                mean=float(np.average(sz, weights=ww)),
                Q25=float(np.interp(0.25, cum, sz)),
                Q75=float(np.interp(0.75, cum, sz)))


def compute_medians(epg_path):
    """Return {well: stats-dict} for every sample lane in the electropherogram."""
    df = _read_table(epg_path)
    # ladder column = the one whose header mentions 'Ladder', else the first column
    lad_col = next((c for c in df.columns if "ladder" in str(c).lower()), df.columns[0])
    calib, LM_l, UM_l = _calibration(df[lad_col].values.astype(float))
    result = {}
    for col in df.columns:
        if col == lad_col:
            continue
        well = str(col).split(":")[0].strip()           # "B1: 4878" -> "B1"
        st = _stats(df[col].values, calib, LM_l, UM_l)
        if st:
            result[well] = st
    return result


# ----------------------------- file plumbing -------------------------------- #
def _find_inputs(args):
    """Resolve (electropherogram_path, sampletable_path) from CLI args."""
    if len(args) >= 2 and os.path.isfile(args[0]) and os.path.isfile(args[1]):
        return args[0], args[1]
    folder = args[0] if args and os.path.isdir(args[0]) else "."
    epgs = glob.glob(os.path.join(folder, "*[Ee]lectropherogram*.csv"))
    if not epgs:
        sys.exit(f"No *Electropherogram*.csv found in '{folder}'.")
    epg = epgs[0]
    prefix = epg[:epg.lower().rfind("electropherogram")]
    samp = glob.glob(prefix + "*ample*able*.csv")
    samp = samp[0] if samp else None
    return epg, samp


def _style(xlsx_path, median_cols):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    thin = Side(style="thin", color="D9D9D9")
    border = Border(bottom=thin, right=thin)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    new_fill = PatternFill("solid", fgColor="C6E0B4")     # highlight the added column(s)
    for j, cell in enumerate(ws[1], start=1):
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = new_fill if cell.value in median_cols else hdr_fill
        if cell.value in median_cols:
            cell.font = Font(name="Arial", bold=True, color="1F4E78", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 3, 10), 40)
    ws.freeze_panes = "A2"
    wb.save(xlsx_path)


def main(argv):
    full = "--full" in argv
    args = [a for a in argv if a != "--full"]
    epg, samp = _find_inputs(args)
    print(f"Electropherogram : {os.path.basename(epg)}")
    print(f"Sample table     : {os.path.basename(samp) if samp else '(none found)'}")

    stats = compute_medians(epg)
    print(f"Medians computed for {len(stats)} samples.")

    extra = ["Median Size [bp]"] + (["Peak Size [bp]", "Mean Size [bp]",
             "Q25 [bp]", "Q75 [bp]"] if full else [])

    # build well -> sample name map from electropherogram column headers
    epg_df = _read_table(epg)
    well_name_map = {}
    for col in epg_df.columns:
        parts = str(col).split(":", 1)
        if len(parts) == 2:
            well_name_map[parts[0].strip()] = parts[1].strip()

    if samp:
        out = _read_table(samp).copy()
        well_col = next((c for c in out.columns if str(c).strip().lower() == "well"),
                        out.columns[0])
        wells = out[well_col].astype(str).str.strip()
        # ensure Sample Description column exists and is populated
        name_col = next((c for c in out.columns
                         if "sample" in str(c).lower() and "desc" in str(c).lower()), None)
        if name_col is None:
            out.insert(1, "Sample Description",
                       [well_name_map.get(w, "") for w in wells])
            name_col = "Sample Description"
        out["Median Size [bp]"] = [round(stats[w]["median"]) if w in stats else None
                                   for w in wells]
        if full:
            out["Peak Size [bp]"] = [round(stats[w]["mode"]) if w in stats else None for w in wells]
            out["Mean Size [bp]"] = [round(stats[w]["mean"]) if w in stats else None for w in wells]
            out["Q25 [bp]"] = [round(stats[w]["Q25"]) if w in stats else None for w in wells]
            out["Q75 [bp]"] = [round(stats[w]["Q75"]) if w in stats else None for w in wells]
        # reorder: Well, Sample Description first, then remaining columns
        other_cols = [c for c in out.columns if c not in (well_col, name_col)]
        out = out[[well_col, name_col] + other_cols]
        base = samp[:samp.lower().rfind("sample")] if "sample" in samp.lower() else os.path.splitext(samp)[0] + "_"
    else:
        rows = [{"Well": w, "Sample Description": well_name_map.get(w, ""),
                 "Median Size [bp]": round(s["median"]),
                 "Peak Size [bp]": round(s["mode"]), "Mean Size [bp]": round(s["mean"]),
                 "Q25 [bp]": round(s["Q25"]), "Q75 [bp]": round(s["Q75"])}
                for w, s in stats.items()]
        out = pd.DataFrame(rows)
        extra = list(out.columns)[2:]
        base = epg[:epg.lower().rfind("electropherogram")]

    out_path = base + "withMedian.xlsx"
    out.to_excel(out_path, index=False)
    _style(out_path, set(extra))
    print(f"\nSaved -> {out_path}")


if __name__ == "__main__":
    main(sys.argv[1:])
